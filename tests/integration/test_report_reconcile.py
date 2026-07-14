from __future__ import annotations

import unittest
from collections.abc import Callable
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from hpr.errors import ReconciliationError
from hpr.excel.build_pivots import PivotAutomationResult
from hpr.report import publish
from hpr.report.build import report_file_name
from hpr.report.run import run_report
from tests.fixtures.report_data import (
    hosted_row,
    reactivated_row,
    write_hosted_csv,
    write_prior_workbook,
    write_reactivated_csv,
)


class ReconciliationTests(unittest.TestCase):
    def test_market_detail_row_count_mismatch_blocks_publish_and_preserves_previous_final(self) -> None:
        self._assert_reconciliation_failure_preserves_previous(
            lambda workbook: workbook["Baton Rouge"].delete_rows(2),
            "Market detail row count mismatch",
        )

    def test_copy_sheet_uid_count_mismatch_blocks_publish_and_preserves_previous_final(self) -> None:
        self._assert_reconciliation_failure_preserves_previous(
            lambda workbook: workbook["Copy"].delete_rows(2),
            "Copy sheet UID count mismatch",
        )

    def test_summary_grand_total_guest_mismatch_blocks_publish_and_preserves_previous_final(self) -> None:
        def mutate(workbook) -> None:
            row_index = _grand_total_row(workbook["Summary"])
            workbook["Summary"].cell(row_index, 3, value=999)

        self._assert_reconciliation_failure_preserves_previous(
            mutate,
            "Summary Grand Total guest count mismatch",
        )

    def test_summary_grand_total_theo_mismatch_blocks_publish_and_preserves_previous_final(self) -> None:
        def mutate(workbook) -> None:
            row_index = _grand_total_row(workbook["Summary"])
            workbook["Summary"].cell(row_index, 4, value=999)

        self._assert_reconciliation_failure_preserves_previous(
            mutate,
            "Summary Grand Total Total Theo mismatch",
        )

    def test_wrong_market_reactivated_file_blocks_publish_and_preserves_previous_final(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            final = root / report_file_name("Baton Rouge", date(2026, 6, 1))
            _write_marker_workbook(final, "old")
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])
            reactivated_csv = write_reactivated_csv(
                root / "reactivated.csv",
                [reactivated_row(**{"Property ID": "HTO", "Universal Player ID": "999"})],
            )

            with self.assertRaisesRegex(ReconciliationError, "outside Baton Rouge"):
                run_report(
                    hosted_csv_path=hosted_csv,
                    last_week_xlsx_path=prior,
                    market="Baton Rouge",
                    report_date=date(2026, 6, 1),
                    output_folder=root,
                    reactivated_csv_path=reactivated_csv,
                    create_native_pivot=False,
                    overwrite=True,
                )

            self.assertEqual(_read_marker(final), "old")

    def test_required_native_pivot_mismatch_blocks_publish_and_preserves_previous_final(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            final = root / report_file_name("Baton Rouge", date(2026, 6, 1))
            _write_marker_workbook(final, "old")
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])
            pivot_result = PivotAutomationResult(
                True,
                "reported created but summary flag is false",
                summary_created=False,
                reactivation_created=False,
            )

            with (
                patch("hpr.report.run.create_native_summary_pivot", return_value=pivot_result),
                self.assertRaisesRegex(ReconciliationError, "Required native PivotTable"),
            ):
                run_report(
                    hosted_csv_path=hosted_csv,
                    last_week_xlsx_path=prior,
                    market="Baton Rouge",
                    report_date=date(2026, 6, 1),
                    output_folder=root,
                    create_native_pivot=True,
                    require_native_pivot=True,
                    overwrite=True,
                )

            self.assertEqual(_read_marker(final), "old")

    def test_reactivated_uid_subset_warning_still_publishes(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])
            reactivated_csv = write_reactivated_csv(
                root / "reactivated.csv",
                [reactivated_row(**{"Universal Player ID": "888"})],
            )

            with self.assertLogs("hpr.report.publish", level="WARNING") as caught:
                result = run_report(
                    hosted_csv_path=hosted_csv,
                    last_week_xlsx_path=prior,
                    market="Baton Rouge",
                    report_date=date(2026, 6, 1),
                    output_folder=root,
                    reactivated_csv_path=reactivated_csv,
                    create_native_pivot=False,
                    overwrite=True,
                )

            self.assertTrue(result.workbook_path.exists())
            self.assertIn("passed_with_warnings", result.reconciliation_status)
            self.assertTrue(any("reactivated_uid_not_subset" in message for message in caught.output))

    def test_clean_reconciliation_publishes_without_warning(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])
            reactivated_csv = write_reactivated_csv(
                root / "reactivated.csv",
                [reactivated_row(**{"Universal Player ID": "999"})],
            )

            with self.assertNoLogs("hpr.report.publish", level="WARNING"):
                result = run_report(
                    hosted_csv_path=hosted_csv,
                    last_week_xlsx_path=prior,
                    market="Baton Rouge",
                    report_date=date(2026, 6, 1),
                    output_folder=root,
                    reactivated_csv_path=reactivated_csv,
                    create_native_pivot=False,
                    overwrite=True,
                )

            self.assertEqual(result.reconciliation_status, "passed")

    def test_static_guest_count_matches_nonblank_uid_reconciliation_count(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            hosted_csv = write_hosted_csv(
                root / "hosted.csv",
                [
                    hosted_row(),
                    hosted_row(
                        **{
                            "Universal Player ID": "",
                            "Guest ID": 2,
                            "Name Full": "Blank UID",
                        }
                    ),
                ],
            )
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])

            result = run_report(
                hosted_csv_path=hosted_csv,
                last_week_xlsx_path=prior,
                market="Baton Rouge",
                report_date=date(2026, 6, 1),
                output_folder=root,
                create_native_pivot=False,
                overwrite=True,
            )

            workbook = load_workbook(result.workbook_path, data_only=True)
            try:
                row_index = _grand_total_row(workbook["Summary"])
                self.assertEqual(workbook["Summary"].cell(row_index, 3).value, 1)
            finally:
                workbook.close()
            self.assertEqual(result.reconciliation_status, "passed")

    def _assert_reconciliation_failure_preserves_previous(
        self,
        mutate: Callable[[Workbook], None],
        message: str,
    ) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            final = root / report_file_name("Baton Rouge", date(2026, 6, 1))
            _write_marker_workbook(final, "old")
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])
            original_save = publish.save_workbook

            def save_then_mutate(workbook: Workbook, workbook_path: Path) -> None:
                original_save(workbook, workbook_path)
                staged = load_workbook(workbook_path)
                try:
                    mutate(staged)
                    staged.save(workbook_path)
                finally:
                    staged.close()

            with (
                patch("hpr.report.publish.save_workbook", side_effect=save_then_mutate),
                self.assertRaisesRegex(ReconciliationError, message),
            ):
                run_report(
                    hosted_csv_path=hosted_csv,
                    last_week_xlsx_path=prior,
                    market="Baton Rouge",
                    report_date=date(2026, 6, 1),
                    output_folder=root,
                    create_native_pivot=False,
                    overwrite=True,
                )

            self.assertEqual(_read_marker(final), "old")


def _write_marker_workbook(path: Path, marker: str) -> None:
    workbook = Workbook()
    workbook.active.title = "Previous"
    workbook.active["A1"] = marker
    workbook.save(path)
    workbook.close()


def _read_marker(path: Path) -> str:
    workbook = load_workbook(path, read_only=True)
    try:
        return workbook["Previous"]["A1"].value
    finally:
        workbook.close()


def _grand_total_row(worksheet) -> int:
    for row_index in range(1, worksheet.max_row + 1):
        if worksheet.cell(row_index, 2).value == "Grand Total":
            return row_index
    raise AssertionError("Grand Total row not found")


if __name__ == "__main__":
    unittest.main()
