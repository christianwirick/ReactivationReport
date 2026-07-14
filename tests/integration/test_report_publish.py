from __future__ import annotations

import unittest
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from hpr.errors import ExcelAutomationError, WorkbookBuildError
from hpr.excel.build_pivots import PivotAutomationResult
from hpr.report.build import report_file_name
from hpr.report.publish import create_staged_workbook_path, remove_stage_file
from hpr.report.run import run_report
from tests.fixtures.report_data import (
    hosted_row,
    reactivated_row,
    write_hosted_csv,
    write_prior_workbook,
    write_reactivated_csv,
)


class WorkbookPublicationTests(unittest.TestCase):
    def test_staged_workbook_path_is_not_the_published_path(self) -> None:
        with TemporaryDirectory() as temp:
            final = Path(temp) / "report.xlsx"
            staged = create_staged_workbook_path(final)
            try:
                self.assertNotEqual(staged, final)
                self.assertEqual(staged.parent, final.parent)
            finally:
                remove_stage_file(staged)

    def test_failed_staged_validation_preserves_previous_final_workbook(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            final = root / report_file_name("Baton Rouge", date(2026, 6, 1))
            _write_marker_workbook(final, "old")
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])

            with (
                patch(
                    "hpr.report.publish.validate_saved_workbook_tabs",
                    side_effect=WorkbookBuildError("forced validation failure"),
                ),
                self.assertRaises(WorkbookBuildError),
            ):
                run_report(
                    hosted_csv_path=hosted_csv,
                    last_week_xlsx_path=prior,
                    market="Baton Rouge",
                    report_date=date(2026, 6, 1),
                    output_folder=root,
                    create_native_pivot=False,
                    overwrite=True,
                )

            self.assertEqual(_read_marker(final), "old")

    def test_optional_com_failure_publishes_valid_static_workbook(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])
            reactivated_csv = write_reactivated_csv(root / "reactivated.csv", [reactivated_row()])
            pivot_result = PivotAutomationResult(False, "Skipped native PivotTable: forced failure.")

            with patch("hpr.report.run.create_native_summary_pivot", return_value=pivot_result):
                result = run_report(
                    hosted_csv_path=hosted_csv,
                    last_week_xlsx_path=prior,
                    market="Baton Rouge",
                    report_date=date(2026, 6, 1),
                    output_folder=root,
                    reactivated_csv_path=reactivated_csv,
                    create_native_pivot=True,
                    require_native_pivot=False,
                    overwrite=True,
                )

            self.assertTrue(result.workbook_path.exists())
            self.assertFalse(result.pivot_result.created)

    def test_required_com_failure_preserves_previous_final_workbook(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            final = root / report_file_name("Baton Rouge", date(2026, 6, 1))
            _write_marker_workbook(final, "old")
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["999"])
            pivot_result = PivotAutomationResult(False, "Skipped native PivotTable: forced failure.")

            with (
                patch("hpr.report.run.create_native_summary_pivot", return_value=pivot_result),
                self.assertRaises(ExcelAutomationError),
            ):
                run_report(
                    hosted_csv_path=hosted_csv,
                    last_week_xlsx_path=prior,
                    market="Baton Rouge",
                    report_date=date(2026, 6, 1),
                    output_folder=root,
                    create_native_pivot=True,
                    require_native_pivot=True,
                    overwrite=True,
                )

            self.assertEqual(_read_marker(final), "old")


def _write_marker_workbook(path: Path, marker: str) -> None:
    workbook = Workbook()
    workbook.active.title = "Previous"
    workbook.active["A1"] = marker
    workbook.save(path)
    workbook.close()


def _read_marker(path: Path) -> str:
    workbook = load_workbook(path, read_only=True)
    try:
        return workbook["Previous"]["A1"].value
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
