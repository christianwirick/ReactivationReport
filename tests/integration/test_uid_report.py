from __future__ import annotations

import unittest
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from hpr.report.run import run_report, run_uid_handoff
from tests.fixtures.report_data import (
    hosted_row,
    reactivated_row,
    workbook_sheet_names,
    write_hosted_csv,
    write_prior_workbook,
    write_reactivated_csv,
)


class UidPipelineTests(unittest.TestCase):
    def test_zero_missing_uids_is_valid_for_handoff_and_report(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row()])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["100"])
            reactivated_csv = write_reactivated_csv(root / "reactivated.csv", [reactivated_row()])

            handoff = run_uid_handoff(
                hosted_csv_path=hosted_csv,
                last_week_xlsx_path=prior,
                market="LBR",
                report_date=date(2026, 6, 1),
                output_folder=root,
            )

            self.assertEqual(handoff.missing_prior_row_count, 0)
            self.assertEqual(handoff.distinct_missing_uid_count, 0)
            self.assertEqual(handoff.clipboard_uids, ())
            self.assertGreaterEqual(handoff.missing_prior_row_count, handoff.distinct_missing_uid_count)

            result = run_report(
                hosted_csv_path=hosted_csv,
                last_week_xlsx_path=prior,
                market="LBR",
                report_date=date(2026, 6, 1),
                output_folder=root,
                reactivated_csv_path=reactivated_csv,
                create_native_pivot=False,
                overwrite=True,
            )

            self.assertTrue(result.workbook_path.exists())
            self.assertEqual(result.missing_prior_row_count, 0)
            self.assertEqual(result.distinct_missing_uid_count, 0)
            self.assertGreaterEqual(result.missing_prior_row_count, result.distinct_missing_uid_count)
            self.assertNotIn("Copy", workbook_sheet_names(result.workbook_path))
            self.assertTrue(result.reconciliation_status.startswith("passed"))

    def test_qa_off_by_default_omits_copy_sheet_and_reconciliation_passes(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            hosted_csv = write_hosted_csv(
                root / "hosted.csv",
                [hosted_row(**{"Universal Player ID": "200"})],
            )
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["100"])

            result = run_report(
                hosted_csv_path=hosted_csv,
                last_week_xlsx_path=prior,
                market="Baton Rouge",
                report_date=date(2026, 6, 1),
                output_folder=root,
                create_native_pivot=False,
                overwrite=True,
            )

            self.assertNotIn("Copy", workbook_sheet_names(result.workbook_path))
            self.assertEqual(result.reconciliation_status, "passed")

    def test_qa_on_adds_hidden_copy_sheet_with_missing_uid_rows(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            hosted_csv = write_hosted_csv(
                root / "hosted.csv",
                [hosted_row(**{"Universal Player ID": "200"})],
            )
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["100", "100.0", "200"])

            result = run_report(
                hosted_csv_path=hosted_csv,
                last_week_xlsx_path=prior,
                market="Baton Rouge",
                report_date=date(2026, 6, 1),
                output_folder=root,
                create_native_pivot=False,
                include_copy_sheet=True,
                overwrite=True,
            )

            workbook = load_workbook(result.workbook_path, data_only=False)
            try:
                copy_sheet = workbook["Copy"]
                self.assertEqual(copy_sheet.sheet_state, "hidden")
                self.assertEqual(
                    [copy_sheet.cell(row_index, 1).value for row_index in range(2, copy_sheet.max_row + 1)],
                    ["100", "100"],
                )
            finally:
                workbook.close()
            self.assertEqual(result.reconciliation_status, "passed")

    def test_duplicate_missing_uids_preserve_clipboard_rows_and_count_distinct(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            hosted_csv = write_hosted_csv(root / "hosted.csv", [hosted_row(**{"Universal Player ID": "200"})])
            prior = write_prior_workbook(root / "prior.xlsx", "Baton Rouge", ["100", "100.0", "200"])

            handoff = run_uid_handoff(
                hosted_csv_path=hosted_csv,
                last_week_xlsx_path=prior,
                market="Baton Rouge",
                report_date=date(2026, 6, 1),
                output_folder=root,
            )

            self.assertEqual(handoff.missing_uids, ("100", "100"))
            self.assertEqual(handoff.clipboard_uids, ("100", "100"))
            self.assertEqual(handoff.missing_prior_row_count, 2)
            self.assertEqual(handoff.distinct_missing_uid_count, 1)
            self.assertGreaterEqual(handoff.missing_prior_row_count, handoff.distinct_missing_uid_count)


if __name__ == "__main__":
    unittest.main()
