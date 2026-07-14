"""Build report inputs for tests."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from hpr.schema import HOSTED_COLUMNS, REACTIVATED_COLUMNS


def hosted_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "Property ID": "LBR",
        "Current Host Name": "Host A",
        "Universal Player ID": "100",
        "Guest ID": 1,
        "Name Full": "Guest One",
        "Tier": "Gold",
        "Trips in Last 90 Days": 1,
        "Last Rated Day": date(2026, 6, 1),
        "Days Since Last Visit": 3,
        "Normal Days In Between Visits": 9,
        "Visit Var": -6,
        "Last 90 ADT": 12.5,
        "Total Theo ": 50,
    }
    row.update(overrides)
    return row


def reactivated_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "Property ID": "LBR",
        "Current Host Name": "Host A",
        "Universal Player ID": "100",
        "Guest ID": 1,
        "Name Full": "Guest One",
        "Tier": "Gold",
        "Last Rated Day": date(2026, 6, 1),
        "Slot Promo Cash In Amt": 5,
        "Slot Theo": 10,
        "Table Theo": 0,
        "Sportsbook Theo": 0,
        "Total Theo": 10,
    }
    row.update(overrides)
    return row


def write_tableau_tsv(path: Path, columns: list[str], rows: list[dict[str, object]]) -> Path:
    lines = ["\t".join(columns)]
    for row in rows:
        values = [_format_tsv_value(row.get(column)) for column in columns]
        lines.append("\t".join(values))
    path.write_text("\n".join(lines) + "\n", encoding="utf-16")
    return path


def write_hosted_csv(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_tableau_tsv(path, HOSTED_COLUMNS, rows)


def write_reactivated_csv(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_tableau_tsv(path, REACTIVATED_COLUMNS, rows)


def write_prior_workbook(path: Path, sheet_name: str, uids: list[object]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["Universal Player ID"])
    for uid in uids:
        worksheet.append([uid])
    workbook.save(path)
    workbook.close()
    return path


def workbook_sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _format_tsv_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, date):
        return f"{value:%m/%d/%Y}"
    return str(value)
