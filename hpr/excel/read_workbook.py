"""Read prior workbooks and Excel theme data."""

from __future__ import annotations

import logging
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from ..diff_uid import normalize_uid
from ..errors import InputValidationError, WorkbookAccessError

logger = logging.getLogger(__name__)


def read_uids_from_workbook_candidates(path: str | Path, sheet_names: list[str]) -> list[str]:
    """Read Universal Player IDs from the first matching candidate sheet."""

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise InputValidationError(f"File does not exist: {workbook_path}")

    try:
        logger.info(
            "input_file_opened operation=prior_workbook_read source_file=%s",
            workbook_path.name,
        )
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    except PermissionError as exc:
        raise WorkbookAccessError(
            "Could not open workbook because Windows denied access: "
            f"{workbook_path.name}. Close the file in Excel, make sure OneDrive has "
            "downloaded it locally, or copy it to a folder you can access and try again."
        ) from exc
    except Exception as exc:  # openpyxl raises several parser exceptions.
        raise WorkbookAccessError(f"Could not open workbook: {workbook_path.name}") from exc

    sheet_name = next((candidate for candidate in sheet_names if candidate in workbook.sheetnames), None)
    if sheet_name is None:
        available = ", ".join(workbook.sheetnames)
        expected = ", ".join(sheet_names)
        raise WorkbookAccessError(
            f"{workbook_path.name} does not contain one of these sheet(s): {expected}. Available sheets: {available}"
        )

    try:
        worksheet = workbook[sheet_name]
        logger.info(
            "prior_workbook_sheet_selected operation=prior_workbook_read source_file=%s stage=sheet_selection sheet=%s",
            workbook_path.name,
            sheet_name,
        )
        header_row, uid_column = _find_header_column(worksheet, "Universal Player ID")
        uids: list[str] = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            value = row[uid_column - 1] if len(row) >= uid_column else None
            uid = normalize_uid(value)
            if uid:
                uids.append(uid)
        logger.info(
            "prior_workbook_uids_read operation=prior_workbook_read source_file=%s sheet=%s row_count=%s result=success",
            workbook_path.name,
            sheet_name,
            len(uids),
        )
        return uids
    finally:
        workbook.close()


def extract_theme_xml(theme_path: str | Path) -> bytes | None:
    """Extract Office theme XML from a .thmx file, if one is available."""

    path = Path(theme_path)
    if not path.exists():
        return None

    try:
        with zipfile.ZipFile(path) as archive:
            return archive.read("theme/theme/theme1.xml")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise InputValidationError(f"Could not read Excel theme file: {path.name}") from exc
    except OSError as exc:
        raise InputValidationError(f"Could not open Excel theme file: {path}. {exc}") from exc


def _find_header_column(worksheet: ReadOnlyWorksheet, header_name: str) -> tuple[int, int]:
    target = header_name.strip().lower()
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        for col_index, value in enumerate(row, start=1):
            if str(value or "").strip().lower() == target:
                return row_index, col_index
    raise InputValidationError(f"Sheet '{worksheet.title}' does not contain a '{header_name}' column.")
