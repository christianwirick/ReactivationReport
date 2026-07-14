"""Build PivotTables with Excel."""

from __future__ import annotations

import logging
import platform
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl.utils import get_column_letter

from ..schema import EXCEL_CURRENCY_FORMAT, SUMMARY_TITLE_FONT, summary_date_label

logger = logging.getLogger(__name__)

XL_DATABASE = 1
XL_ROW_FIELD = 1
XL_COUNT = -4112
XL_SUM = -4157
XL_UP = -4162
XL_TO_LEFT = -4159
XL_A1 = 1
XL_R1C1 = -4150


@dataclass(frozen=True)
class PivotAutomationResult:
    created: bool
    message: str
    theme_message: str = ""
    summary_created: bool = False
    reactivation_created: bool = False


def create_native_summary_pivot(
    *,
    workbook_path: str | Path,
    market_sheet: str,
    report_date: date,
    include_reactivation_summary: bool = False,
    theme_path: str | Path | None = None,
) -> PivotAutomationResult:
    """Replace static summary sheets with native Excel PivotTables."""

    if platform.system() != "Windows":
        logger.warning(
            "native_pivot_status operation=excel_automation result=skipped fallback_state=static_summary reason=non_windows"
        )
        return PivotAutomationResult(
            False, "Skipped native PivotTable: Windows Excel automation is unavailable on this OS."
        )

    try:
        import pythoncom
        import win32com.client as win32
    except ImportError:
        logger.warning(
            "native_pivot_status operation=excel_automation result=skipped fallback_state=static_summary reason=pywin32_missing"
        )
        return PivotAutomationResult(False, "Skipped native PivotTable: pywin32 is not installed.")

    workbook = None
    excel = None
    pythoncom.CoInitialize()
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.EnableEvents = False
            excel.ScreenUpdating = False
            excel.AskToUpdateLinks = False
        except Exception:
            logger.debug("Could not disable every Excel prompt/update setting", exc_info=True)

        workbook_path = Path(workbook_path).resolve()
        workbook = excel.Workbooks.Open(str(workbook_path), UpdateLinks=0, ReadOnly=False, Notify=False)
        theme_message = _apply_theme_if_available(workbook, theme_path)

        if not _worksheet_exists(workbook, market_sheet):
            return PivotAutomationResult(
                False,
                f"Skipped native PivotTable: workbook does not contain '{market_sheet}'.",
            )

        messages: list[str] = []
        summary_created = False
        reactivation_created = False

        data_ws = workbook.Worksheets(market_sheet)
        try:
            summary_ws = _prepare_existing_summary_sheet(workbook, "Summary")
            _create_summary_pivot(
                workbook=workbook,
                source_ws=data_ws,
                summary_ws=summary_ws,
                table_name="HostedPlayersPivot",
                report_date=report_date,
            )
            _assert_native_pivot_exists(summary_ws, "Summary")
            _move_sheet_first(workbook, summary_ws)
            summary_created = True
            messages.append("Summary native PivotTable created")
        except Exception as exc:
            messages.append(f"Summary native PivotTable failed: {exc}")

        if include_reactivation_summary and _worksheet_exists(workbook, "Reactivated Players"):
            try:
                reactivated_ws = workbook.Worksheets("Reactivated Players")
                reactivation_summary_ws = _prepare_existing_summary_sheet(workbook, "Reactivation Summary")
                _move_sheet_after(reactivation_summary_ws, reactivated_ws)
                _create_reactivation_pivot(
                    workbook=workbook,
                    source_ws=reactivated_ws,
                    summary_ws=reactivation_summary_ws,
                    table_name="ReactivatedPlayersPivot",
                    report_date=report_date,
                )
                _assert_native_pivot_exists(reactivation_summary_ws, "Reactivation Summary")
                reactivation_created = True
                messages.append("Reactivation Summary native PivotTable created")
            except Exception as exc:
                messages.append(f"Reactivation Summary native PivotTable failed: {exc}")
        elif include_reactivation_summary:
            messages.append("Reactivation Summary native PivotTable failed: Reactivated Players sheet was not found")

        all_requested_created = summary_created and (reactivation_created or not include_reactivation_summary)
        if not all_requested_created:
            _close_workbook(workbook, save_changes=False, context="degraded_cleanup")
            workbook = None
            logger.warning(
                "native_pivot_status operation=excel_automation result=degraded fallback_state=static_summary target_path=%s message=%s",
                workbook_path.name,
                "; ".join(messages),
            )
            return PivotAutomationResult(
                False,
                "; ".join(messages),
                theme_message,
                summary_created,
                reactivation_created,
            )

        workbook.Worksheets("Summary").Activate()
        excel.ActiveWindow.DisplayGridlines = False

        workbook.Save()
        _verify_final_native_sheets(
            workbook=workbook,
            include_reactivation_summary=include_reactivation_summary,
        )
        _close_workbook(workbook, save_changes=True, context="success_cleanup")
        workbook = None
        logger.info(
            "native_pivot_status operation=excel_automation result=success target_path=%s",
            workbook_path.name,
        )
        return PivotAutomationResult(
            True,
            "; ".join(messages),
            theme_message,
            summary_created,
            reactivation_created,
        )
    except Exception as exc:
        if workbook is not None:
            _close_workbook(workbook, save_changes=False, context="error_cleanup")
            workbook = None
        logger.warning(
            "native_pivot_status operation=excel_automation result=degraded fallback_state=static_summary error_type=%s",
            type(exc).__name__,
        )
        return PivotAutomationResult(False, f"Skipped native PivotTable after Excel automation error: {exc}")
    finally:
        if workbook is not None:
            _close_workbook(workbook, save_changes=False, context="finally_cleanup")
        if excel is not None:
            _quit_excel(excel)
        _co_uninitialize(pythoncom)


def _close_workbook(workbook, *, save_changes: bool, context: str) -> None:
    try:
        workbook.Close(SaveChanges=save_changes)
    except Exception:
        logger.warning(
            "excel_teardown_failed operation=excel_automation stage=workbook_close context=%s",
            context,
            exc_info=True,
        )


def _quit_excel(excel) -> None:
    try:
        excel.Quit()
    except Exception:
        logger.warning(
            "excel_teardown_failed operation=excel_automation stage=excel_quit",
            exc_info=True,
        )


def _co_uninitialize(pythoncom) -> None:
    try:
        pythoncom.CoUninitialize()
    except Exception:
        logger.warning(
            "excel_teardown_failed operation=excel_automation stage=co_uninitialize",
            exc_info=True,
        )


def _worksheet_exists(workbook, sheet_name: str) -> bool:
    for index in range(1, workbook.Worksheets.Count + 1):
        worksheet = workbook.Worksheets(index)
        if worksheet.Name == sheet_name:
            return True
    return False


def _apply_theme_if_available(workbook, theme_path: str | Path | None) -> str:
    if not theme_path:
        return "Workbook theme not applied: no theme path was provided."
    path = Path(theme_path)
    if not path.exists():
        return f"Workbook theme not applied: file not found at {path}."
    if path.suffix.lower() != ".thmx":
        return f"Workbook theme not applied: expected a .thmx file, got {path}."
    try:
        workbook.ApplyTheme(str(path.resolve()))
        return f"Workbook theme applied through Excel: {path.name}"
    except Exception as first_exc:
        try:
            workbook.Application.ActiveWorkbook.ApplyTheme(str(path.resolve()))
            return f"Workbook theme applied through Excel: {path.name}"
        except Exception as second_exc:
            return (
                "Workbook theme failed to apply through Excel. "
                f"Workbook.ApplyTheme error: {first_exc}; ActiveWorkbook.ApplyTheme error: {second_exc}"
            )


def _prepare_existing_summary_sheet(workbook, sheet_name: str):
    if _worksheet_exists(workbook, sheet_name):
        worksheet = workbook.Worksheets(sheet_name)
    else:
        worksheet = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
        worksheet.Name = sheet_name
    worksheet.Cells.Clear()
    return worksheet


def _move_sheet_first(workbook, worksheet) -> None:
    if worksheet.Index != 1:
        worksheet.Move(Before=workbook.Worksheets(1))


def _move_sheet_after(worksheet, after_worksheet) -> None:
    if worksheet.Index != after_worksheet.Index + 1:
        worksheet.Move(After=after_worksheet)


def _verify_final_native_sheets(*, workbook, include_reactivation_summary: bool) -> None:
    _assert_native_pivot_exists(workbook.Worksheets("Summary"), "Summary")
    if include_reactivation_summary:
        if not _worksheet_exists(workbook, "Reactivation Summary"):
            raise RuntimeError("Reactivation Summary sheet was not present after saving.")
        _assert_native_pivot_exists(workbook.Worksheets("Reactivation Summary"), "Reactivation Summary")


def _assert_native_pivot_exists(worksheet, label: str) -> None:
    try:
        count = worksheet.PivotTables().Count
    except Exception as exc:
        raise RuntimeError(f"{label} sheet did not expose PivotTables().Count") from exc
    if count < 1:
        raise RuntimeError(f"{label} sheet has no native PivotTable after creation")


def _create_summary_pivot(*, workbook, source_ws, summary_ws, table_name: str, report_date: date) -> None:
    pivot = _create_pivot(
        workbook=workbook,
        source_ws=source_ws,
        destination_ws=summary_ws,
        table_name=table_name,
        required_field_groups=[
            ["Property ID"],
            ["Current Host Name"],
            ["Universal Player ID"],
            ["Total Theo ", "Total Theo"],
        ],
    )
    _add_row_field(pivot, ["Property ID"], position=1)
    _add_row_field(pivot, ["Current Host Name"], position=2)
    _add_count_field(pivot, ["Universal Player ID"], "Guests")
    _add_sum_field(pivot, ["Total Theo ", "Total Theo"], "Sum of Total Theo")
    _decorate_summary_sheet(summary_ws, "Summary", report_date, last_column="D")


def _create_reactivation_pivot(*, workbook, source_ws, summary_ws, table_name: str, report_date: date) -> None:
    pivot = _create_pivot(
        workbook=workbook,
        source_ws=source_ws,
        destination_ws=summary_ws,
        table_name=table_name,
        required_field_groups=[
            ["Property ID"],
            ["Current Host Name"],
            ["Universal Player ID"],
            ["Slot Promo Cash In Amt"],
            ["Slot Theo"],
            ["Table Theo"],
            ["Sportsbook Theo"],
            ["Total Theo"],
        ],
    )
    _add_row_field(pivot, ["Property ID"], position=1)
    _add_row_field(pivot, ["Current Host Name"], position=2)
    _add_count_field(pivot, ["Universal Player ID"], "Guests")
    _add_sum_field(pivot, ["Slot Promo Cash In Amt"], "Sum of Slot Promo Cash In Amt")
    _add_sum_field(pivot, ["Slot Theo"], "Sum of Slot Theo")
    _add_sum_field(pivot, ["Table Theo"], "Sum of Table Theo")
    _add_sum_field(pivot, ["Sportsbook Theo"], "Sum of Sportsbook Theo")
    _add_sum_field(pivot, ["Total Theo"], "Sum of Total Theo")
    _decorate_summary_sheet(summary_ws, "Reactivation Summary", report_date, last_column="H")


def _create_pivot(
    *,
    workbook,
    source_ws,
    destination_ws,
    table_name: str,
    required_field_groups: list[list[str]],
):
    last_row = source_ws.Cells(source_ws.Rows.Count, 1).End(XL_UP).Row
    last_col = source_ws.Cells(1, source_ws.Columns.Count).End(XL_TO_LEFT).Column
    _validate_pivot_source(
        source_ws=source_ws,
        last_row=last_row,
        last_col=last_col,
        required_field_groups=required_field_groups,
    )
    source_range = source_ws.Range(source_ws.Cells(1, 1), source_ws.Cells(last_row, last_col))
    pivot_cache = _create_pivot_cache(
        workbook=workbook,
        source_range=source_range,
        sheet_name=source_ws.Name,
        last_row=last_row,
        last_col=last_col,
    )
    pivot = pivot_cache.CreatePivotTable(destination_ws.Range("B4"), table_name)
    pivot.RefreshTable()
    return pivot


def _validate_pivot_source(
    *,
    source_ws,
    last_row: int,
    last_col: int,
    required_field_groups: list[list[str]],
) -> None:
    if last_row < 2:
        raise RuntimeError(f"'{source_ws.Name}' has no data rows for the PivotTable.")
    if last_col < 1:
        raise RuntimeError(f"'{source_ws.Name}' has no columns for the PivotTable.")

    headers = [str(source_ws.Cells(1, column_index).Value or "").strip() for column_index in range(1, last_col + 1)]
    if any(not header for header in headers):
        blank_columns = [get_column_letter(index) for index, header in enumerate(headers, start=1) if not header]
        raise RuntimeError(
            f"'{source_ws.Name}' has blank header cell(s): {', '.join(blank_columns)}. "
            "Excel cannot build a reliable PivotTable from this range."
        )

    normalized_headers = {_normalize_field_name(header) for header in headers}
    missing = [
        " / ".join(group)
        for group in required_field_groups
        if not any(_normalize_field_name(candidate) in normalized_headers for candidate in group)
    ]
    if missing:
        raise RuntimeError(
            f"'{source_ws.Name}' is missing PivotTable field(s): {', '.join(missing)}. "
            f"Available fields: {', '.join(headers)}"
        )


def _create_pivot_cache(*, workbook, source_range, sheet_name: str, last_row: int, last_col: int):
    attempts = []
    sources = [
        ("Excel Range object", lambda: source_range),
        (
            "external R1C1 address",
            lambda: _external_r1c1_range_address(source_range, sheet_name, last_row, last_col),
        ),
        (
            "external A1 address",
            lambda: _external_range_address(source_range, sheet_name, last_row, last_col),
        ),
        ("quoted R1C1 address", lambda: _quoted_r1c1_range_address(sheet_name, last_row, last_col)),
        ("quoted A1 address", lambda: _quoted_a1_range_address(sheet_name, last_row, last_col)),
    ]

    for label, source_factory in sources:
        try:
            return workbook.PivotCaches().Create(XL_DATABASE, source_factory())
        except Exception as exc:
            attempts.append(f"{label}: {exc}")

    raise RuntimeError("Excel could not create a PivotCache from the source range. Tried " + "; ".join(attempts))


def _external_range_address(source_range, sheet_name: str, last_row: int, last_col: int) -> str:
    """Return an Excel-safe external A1 address for a PivotCache source range."""

    try:
        return source_range.GetAddress(
            RowAbsolute=True,
            ColumnAbsolute=True,
            ReferenceStyle=XL_A1,
            External=True,
        )
    except Exception:
        return _quoted_a1_range_address(sheet_name, last_row, last_col)


def _external_r1c1_range_address(source_range, sheet_name: str, last_row: int, last_col: int) -> str:
    """Return an Excel-safe external R1C1 address for a PivotCache source range."""

    try:
        return source_range.GetAddress(True, True, XL_R1C1, True)
    except Exception:
        return _quoted_r1c1_range_address(sheet_name, last_row, last_col)


def _quoted_a1_range_address(sheet_name: str, last_row: int, last_col: int) -> str:
    quoted_sheet = sheet_name.replace("'", "''")
    end_column = get_column_letter(last_col)
    return f"'{quoted_sheet}'!$A$1:${end_column}${last_row}"


def _quoted_r1c1_range_address(sheet_name: str, last_row: int, last_col: int) -> str:
    quoted_sheet = sheet_name.replace("'", "''")
    return f"'{quoted_sheet}'!R1C1:R{last_row}C{last_col}"


def _add_count_field(pivot, names: list[str], caption: str):
    field = pivot.AddDataField(_pivot_field(pivot, names), caption, XL_COUNT)
    field.NumberFormat = "#,##0"
    return field


def _add_sum_field(pivot, names: list[str], caption: str):
    field = pivot.AddDataField(_pivot_field(pivot, names), caption, XL_SUM)
    field.NumberFormat = EXCEL_CURRENCY_FORMAT
    return field


def _decorate_summary_sheet(summary_ws, title: str, report_date: date, last_column: str) -> None:
    summary_ws.Range("B1").Value = title
    summary_ws.Range("B1").Font.Name = SUMMARY_TITLE_FONT
    summary_ws.Range("B1").Font.Size = 20
    summary_ws.Range("B1").Font.Bold = True

    summary_ws.Range("B2").Value = summary_date_label(report_date)
    summary_ws.Range("B2").Font.Name = SUMMARY_TITLE_FONT
    summary_ws.Range("B2").Font.Size = 14

    summary_ws.Range("B4").Value = "Host Name"
    summary_ws.Range("C4").Value = "Guests"
    summary_ws.Columns(f"B:{last_column}").AutoFit()


def _add_row_field(pivot, names: list[str], position: int) -> None:
    field = _pivot_field(pivot, names)
    field.Orientation = XL_ROW_FIELD
    field.Position = position


def _pivot_field(pivot, names: list[str]):
    last_error = None
    for name in names:
        try:
            return pivot.PivotFields(name)
        except Exception as exc:
            last_error = exc

    target_names = {_normalize_field_name(name) for name in names}
    available_names = []
    try:
        fields = pivot.PivotFields()
        for index in range(1, fields.Count + 1):
            field = fields(index)
            available_names.append(str(field.Name))
            if _normalize_field_name(field.Name) in target_names:
                return field
    except Exception as exc:
        last_error = exc

    available = ", ".join(available_names) if available_names else "unavailable"
    raise RuntimeError(
        f"Could not find PivotTable field: {', '.join(names)}. Available fields: {available}"
    ) from last_error


def _normalize_field_name(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())
