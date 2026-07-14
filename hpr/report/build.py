from __future__ import annotations

import logging
import re
from collections import OrderedDict
from datetime import date
from pathlib import Path
from typing import TypedDict

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from ..diff_uid import normalize_uid
from ..errors import InputValidationError
from ..schema import (
    EXCEL_CURRENCY_FORMAT,
    HOSTED_COLUMN_WIDTHS,
    HOSTED_COLUMNS,
    HOSTED_SUM_COLUMNS,
    REACTIVATED_COLUMN_WIDTHS,
    REACTIVATED_COLUMNS,
    REACTIVATED_SUM_COLUMNS,
    REACTIVATION_SUMMARY_COLUMN_WIDTHS,
    SUMMARY_COLUMN_WIDTHS,
    SUMMARY_TITLE_FONT,
    TEXT_COLUMNS,
    summary_date_label,
)

logger = logging.getLogger(__name__)

WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}
WINDOWS_INVALID_FILENAME_CHARS = set('<>:"/\\|?*')
FORMULA_INJECTION_PREFIXES = ("=", "+", "-", "@")


class SummaryRow(TypedDict):
    name: str
    guests: int
    sums: dict[str, int | float]


def add_hosted_report_sheets(
    *,
    workbook: Workbook,
    market: str,
    report_date: date,
    hosted_rows: list[dict[str, object]],
) -> None:
    """Add summary and detail sheets."""

    logger.info(
        "workbook_stage_started operation=workbook_build stage=hosted_sheets market=%s report_date=%s row_count=%s",
        market,
        report_date.isoformat(),
        len(hosted_rows),
    )
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    _write_summary_sheet(
        summary,
        title="Summary",
        report_date=report_date,
        rows=hosted_rows,
        sum_columns=HOSTED_SUM_COLUMNS,
        widths=SUMMARY_COLUMN_WIDTHS,
    )

    _write_detail_sheet(
        workbook.create_sheet(_safe_sheet_name(market)),
        HOSTED_COLUMNS,
        hosted_rows,
        HOSTED_COLUMN_WIDTHS,
    )

    property_ids = _ordered_values(hosted_rows, "Property ID")
    if len(property_ids) <= 1:
        return

    for property_id in property_ids:
        property_rows = [row for row in hosted_rows if row.get("Property ID") == property_id]
        _write_detail_sheet(
            workbook.create_sheet(_safe_sheet_name(str(property_id), workbook.sheetnames)),
            HOSTED_COLUMNS,
            property_rows,
            HOSTED_COLUMN_WIDTHS,
        )


def report_file_name(market: str, report_date: date) -> str:
    """Return dated report filename."""

    safe_market = safe_filename_component(market)
    return f"{safe_market} - Hosted Players Report {report_date:%m.%d.%y}.xlsx"


def safe_filename_component(value: str) -> str:
    """Validate Windows-safe file name component."""

    text = str(value or "").strip()
    if not text:
        raise InputValidationError("Market name is required for the output filename.")
    if text in {".", ".."}:
        raise InputValidationError("Market name cannot be '.' or '..'.")
    if any(char in WINDOWS_INVALID_FILENAME_CHARS for char in text):
        raise InputValidationError("Market name contains characters that are not valid in Windows filenames.")
    if text.rstrip(" .") != text:
        raise InputValidationError("Market name cannot end with a space or period.")
    if text.upper() in WINDOWS_RESERVED_NAMES:
        raise InputValidationError(f"Market name is reserved by Windows: {text}.")
    if Path(text).name != text:
        raise InputValidationError("Market name cannot contain path components.")
    return text


def add_reactivated_report_sheets(
    *,
    workbook: Workbook,
    report_date: date,
    reactivated_rows: list[dict[str, object]],
) -> None:
    """Add reactivated-player summary and detail sheets."""

    logger.info(
        "workbook_stage_started operation=workbook_build stage=reactivated_sheets report_date=%s row_count=%s",
        report_date.isoformat(),
        len(reactivated_rows),
    )
    _delete_sheet_if_exists(workbook, "Reactivated Players")
    _delete_sheet_if_exists(workbook, "Reactivation Summary")

    _write_detail_sheet(
        workbook.create_sheet("Reactivated Players"),
        REACTIVATED_COLUMNS,
        reactivated_rows,
        REACTIVATED_COLUMN_WIDTHS,
    )
    _format_currency_range(workbook["Reactivated Players"], start_col=8, end_col=12)

    reactivation_summary = workbook.create_sheet("Reactivation Summary")
    reactivation_summary.sheet_view.showGridLines = False
    _write_summary_sheet(
        reactivation_summary,
        title="Reactivation Summary",
        report_date=report_date,
        rows=reactivated_rows,
        sum_columns=REACTIVATED_SUM_COLUMNS,
        widths=REACTIVATION_SUMMARY_COLUMN_WIDTHS,
    )


def _write_detail_sheet(
    worksheet: Worksheet,
    columns: list[str],
    rows: list[dict[str, object]],
    widths: dict[str, int],
) -> None:
    worksheet.append(columns)
    for row in rows:
        worksheet.append([_worksheet_value(column, row.get(column)) for column in columns])

    _apply_widths(worksheet, widths)
    _format_date_columns(worksheet, columns)
    _format_currency_columns(worksheet, columns)


def _write_summary_sheet(
    worksheet: Worksheet,
    *,
    title: str,
    report_date: date,
    rows: list[dict[str, object]],
    sum_columns: list[str],
    widths: dict[str, int],
) -> None:
    worksheet.cell(row=1, column=2, value=title)
    worksheet.cell(row=2, column=2, value=summary_date_label(report_date))

    headers = ["Host Name", "Guests"] + [f"Sum of {column}" for column in sum_columns]
    for index, header in enumerate(headers, start=2):
        worksheet.cell(row=4, column=index, value=header)

    current_row = 5
    for summary_row in _summary_rows(rows, sum_columns):
        worksheet.cell(row=current_row, column=2, value=_neutralize_text_cell(summary_row["name"]))
        worksheet.cell(row=current_row, column=3, value=summary_row["guests"])
        for offset, column in enumerate(sum_columns, start=4):
            cell = worksheet.cell(row=current_row, column=offset, value=summary_row["sums"][column])
            if column.strip().endswith("Theo") or column == "Slot Promo Cash In Amt":
                cell.number_format = EXCEL_CURRENCY_FORMAT
        current_row += 1

    worksheet.row_dimensions[1].height = 24.6
    worksheet.row_dimensions[2].height = 17.4
    worksheet["B1"].font = Font(name=SUMMARY_TITLE_FONT, size=20, bold=True, color="000000")
    worksheet["B2"].font = Font(name=SUMMARY_TITLE_FONT, size=14, color="000000")
    _apply_widths(worksheet, widths)


def _summary_rows(rows: list[dict[str, object]], sum_columns: list[str]) -> list[SummaryRow]:
    results: list[SummaryRow] = []
    property_groups: OrderedDict[str, list[dict[str, object]]] = OrderedDict()

    for row in rows:
        property_id = str(row.get("Property ID") or "").strip()
        property_groups.setdefault(property_id, []).append(row)

    for property_id, property_rows in property_groups.items():
        results.append(_aggregate_row(property_id, property_rows, sum_columns))
        host_groups: OrderedDict[str, list[dict[str, object]]] = OrderedDict()
        for row in property_rows:
            host = str(row.get("Current Host Name") or "").strip()
            host_groups.setdefault(host, []).append(row)
        for host, host_rows in host_groups.items():
            results.append(_aggregate_row(host, host_rows, sum_columns))

    results.append(_aggregate_row("Grand Total", rows, sum_columns))
    return results


def _aggregate_row(name: str, rows: list[dict[str, object]], sum_columns: list[str]) -> SummaryRow:
    return {
        "name": name,
        # Match native PivotTable XL_COUNT semantics: non-blank Universal Player ID rows.
        "guests": _guest_count(rows),
        "sums": {column: _sum_column(rows, column) for column in sum_columns},
    }


def _guest_count(rows: list[dict[str, object]]) -> int:
    return sum(1 for row in rows if normalize_uid(row.get("Universal Player ID")))


def _worksheet_value(column: str, value: object) -> object:
    if column in TEXT_COLUMNS:
        return _neutralize_text_cell(value)
    return value


def _neutralize_text_cell(value: object) -> object:
    if isinstance(value, str) and value.startswith(FORMULA_INJECTION_PREFIXES):
        return "'" + value
    return value


def _sum_column(rows: list[dict[str, object]], column: str) -> int | float:
    total: int | float = 0
    for row in rows:
        value = row.get(column)
        if isinstance(value, (int, float)):
            total += value
    return total


def _ordered_values(rows: list[dict[str, object]], column: str) -> list[object]:
    seen: set[object] = set()
    ordered: list[object] = []
    for row in rows:
        value = row.get(column)
        if value not in seen:
            seen.add(value)
            ordered.append(value)
    return ordered


def _apply_widths(worksheet: Worksheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _format_date_columns(worksheet: Worksheet, columns: list[str]) -> None:
    for column_index, header in enumerate(columns, start=1):
        if header != "Last Rated Day":
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = "m/d/yyyy"


def _format_currency_columns(worksheet: Worksheet, columns: list[str]) -> None:
    for column_index, header in enumerate(columns, start=1):
        if header not in {"Last 90 ADT", "Total Theo "}:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = EXCEL_CURRENCY_FORMAT


def _format_currency_range(worksheet: Worksheet, *, start_col: int, end_col: int) -> None:
    for row_index in range(2, worksheet.max_row + 1):
        for column_index in range(start_col, end_col + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = EXCEL_CURRENCY_FORMAT


def _delete_sheet_if_exists(workbook: Workbook, sheet_name: str) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]


def _safe_sheet_name(name: str, existing_names: list[str] | None = None) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", name).strip() or "Sheet"
    cleaned = cleaned[:31]
    if not existing_names or cleaned not in existing_names:
        return cleaned

    base = cleaned[:28]
    index = 2
    while True:
        candidate = f"{base}_{index}"[:31]
        if candidate not in existing_names:
            return candidate
        index += 1
