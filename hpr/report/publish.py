"""Validate/publish staged workbooks."""

from __future__ import annotations

import logging
import os
import tempfile
from collections.abc import Mapping, Sequence
from pathlib import Path
from time import time
from typing import Protocol

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

from ..diff_uid import normalize_uid
from ..errors import (
    OutputPublicationError,
    ReconciliationError,
    WorkbookBuildError,
    WorkbookLockedError,
)
from ..match_market import MARKET_PROPERTY_CODES

logger = logging.getLogger(__name__)


class UidComparisonLike(Protocol):
    @property
    def missing_uids(self) -> tuple[str, ...]: ...

    @property
    def missing_row_count(self) -> int: ...


class PivotResultLike(Protocol):
    @property
    def created(self) -> bool: ...

    @property
    def summary_created(self) -> bool: ...

    @property
    def reactivation_created(self) -> bool: ...


def ensure_can_publish(final_path: Path, *, overwrite: bool) -> None:
    """Confirm output can be published."""

    if final_path.exists() and not overwrite:
        raise OutputPublicationError(
            f"Output already exists: {final_path.name}. Choose a different output folder or pass --overwrite."
        )


def create_staged_workbook_path(final_path: Path) -> Path:
    """Create temporary workbook beside final output."""

    final_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix=f".{final_path.stem}.",
        suffix=".tmp.xlsx",
        dir=final_path.parent,
        delete=False,
    ) as handle:
        return Path(handle.name)


def save_workbook(workbook: Workbook, workbook_path: Path) -> None:
    """Save staged workbook."""

    try:
        workbook.save(workbook_path)
    except PermissionError as exc:
        raise WorkbookLockedError(workbook_path) from exc
    except OSError as exc:
        raise WorkbookBuildError(f"Could not save workbook to {workbook_path}: {exc}") from exc


def validate_saved_workbook_tabs(workbook_path: Path, expected_sheets: list[str]) -> None:
    """Verify all expected sheets exist."""

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    except Exception as exc:
        raise WorkbookBuildError(f"Could not reopen saved workbook for validation: {workbook_path}") from exc

    try:
        missing = [sheet_name for sheet_name in expected_sheets if sheet_name not in workbook.sheetnames]
        if missing:
            raise WorkbookBuildError(
                "Final workbook is missing expected sheet(s): "
                + ", ".join(missing)
                + ". Available sheets: "
                + ", ".join(workbook.sheetnames)
            )
    finally:
        workbook.close()


def reconcile_report(
    *,
    workbook_path: Path,
    market: str,
    hosted_rows: Sequence[Mapping[str, object]],
    reactivated_rows: Sequence[Mapping[str, object]] | None,
    comparison: UidComparisonLike,
    include_copy_sheet: bool,
) -> str:
    """Verify the workbook matches its source data."""

    warnings: list[str] = []
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        raise WorkbookBuildError(f"Could not reopen saved workbook for reconciliation: {workbook_path}") from exc

    try:
        _reconcile_market_detail_count(workbook, market, hosted_rows)
        if include_copy_sheet:
            _reconcile_copy_sheet_count(workbook, comparison)
        _reconcile_summary_grand_total(workbook, hosted_rows)
        if reactivated_rows is not None:
            _reconcile_reactivated_property_codes(market, reactivated_rows)
            warnings.extend(_warn_if_reactivated_uids_not_subset(reactivated_rows, comparison))
    finally:
        workbook.close()

    if warnings:
        return "passed_with_warnings: " + "; ".join(warnings)
    return "passed"


def validate_required_native_pivots(
    *,
    pivot_result: PivotResultLike,
    include_reactivation_summary: bool,
) -> None:
    """Verify PivotTables creatation."""

    if not pivot_result.summary_created:
        raise ReconciliationError("Required native PivotTable was not created for Summary.")
    if include_reactivation_summary and not pivot_result.reactivation_created:
        raise ReconciliationError("Required native PivotTable was not created for Reactivation Summary.")


def publish_workbook(staged_path: Path, final_path: Path, *, overwrite: bool) -> None:
    """Publish the validated workbook atomically."""

    ensure_can_publish(final_path, overwrite=overwrite)
    try:
        os.replace(staged_path, final_path)
    except PermissionError as exc:
        raise WorkbookLockedError(final_path) from exc
    except OSError as exc:
        raise OutputPublicationError(f"Could not publish workbook to {final_path}: {exc}") from exc
    logger.info(
        "output_published operation=workbook_publication target_path=%s result=success",
        final_path.name,
    )


def remove_stage_file(staged_path: Path) -> None:
    """Remove staged workbooks."""

    try:
        if staged_path.exists():
            staged_path.unlink()
    except OSError as exc:
        logger.warning(
            "stage_cleanup_failed operation=workbook_publication target_path=%s error_type=%s",
            staged_path.name,
            type(exc).__name__,
        )


def cleanup_abandoned_stage_files(final_path: Path) -> None:
    """Remove staged workbooks > 1 day old."""

    cutoff = time() - 24 * 60 * 60
    pattern = f".{final_path.stem}.*.tmp.xlsx"
    for path in final_path.parent.glob(pattern):
        try:
            if path.stat().st_mtime >= cutoff:
                continue
            path.unlink()
            logger.info(
                "stage_cleanup_completed operation=workbook_publication target_path=%s result=removed_abandoned",
                path.name,
            )
        except OSError as exc:
            logger.warning(
                "stage_cleanup_failed operation=workbook_publication target_path=%s error_type=%s",
                path.name,
                type(exc).__name__,
            )


def _reconcile_market_detail_count(
    workbook: Workbook, market: str, hosted_rows: Sequence[Mapping[str, object]]
) -> None:
    worksheet = workbook[market]
    actual_rows = max(0, worksheet.max_row - 1)
    expected_rows = len(hosted_rows)
    if actual_rows != expected_rows:
        raise ReconciliationError(
            f"Market detail row count mismatch for {market}: expected {expected_rows}, found {actual_rows}."
        )


def _reconcile_copy_sheet_count(workbook: Workbook, comparison: UidComparisonLike) -> None:
    if "Copy" not in workbook.sheetnames:
        raise ReconciliationError("Hidden Copy sheet is missing.")
    worksheet = workbook["Copy"]
    if worksheet.sheet_state != "hidden":
        raise ReconciliationError("Copy sheet must remain hidden.")
    actual_uids = sum(
        1 for row_index in range(2, worksheet.max_row + 1) if normalize_uid(worksheet.cell(row_index, 1).value)
    )
    expected_uids = comparison.missing_row_count
    if actual_uids != expected_uids:
        raise ReconciliationError(f"Copy sheet UID count mismatch: expected {expected_uids}, found {actual_uids}.")


def _reconcile_summary_grand_total(workbook: Workbook, hosted_rows: Sequence[Mapping[str, object]]) -> None:
    worksheet = workbook["Summary"]
    row_index = _find_grand_total_row(worksheet)
    actual_guests = int(worksheet.cell(row_index, 3).value or 0)
    expected_guests = _nonblank_uid_count(hosted_rows)
    if actual_guests != expected_guests:
        raise ReconciliationError(
            f"Summary Grand Total guest count mismatch: expected {expected_guests}, found {actual_guests}."
        )

    actual_total_cents = _round_to_cents(worksheet.cell(row_index, 4).value)
    expected_total_cents = _round_to_cents(sum(_numeric(row.get("Total Theo ")) for row in hosted_rows))
    if abs(actual_total_cents - expected_total_cents) > 1:
        raise ReconciliationError(
            "Summary Grand Total Total Theo mismatch: "
            f"expected {_format_cents(expected_total_cents)}, found {_format_cents(actual_total_cents)}."
        )


def _reconcile_reactivated_property_codes(market: str, reactivated_rows: Sequence[Mapping[str, object]]) -> None:
    allowed_codes = MARKET_PROPERTY_CODES.get(market, set())
    foreign_codes = sorted(
        {
            str(row.get("Property ID") or "").strip()
            for row in reactivated_rows
            if str(row.get("Property ID") or "").strip() not in allowed_codes
        }
    )
    if foreign_codes:
        raise ReconciliationError(
            f"Reactivated Players CSV contains Property ID code(s) outside {market}: {', '.join(foreign_codes)}."
        )


def _warn_if_reactivated_uids_not_subset(
    reactivated_rows: Sequence[Mapping[str, object]],
    comparison: UidComparisonLike,
) -> list[str]:
    missing_uids = {normalize_uid(uid) for uid in comparison.missing_uids if normalize_uid(uid)}
    reactivated_uids = {
        normalize_uid(row.get("Universal Player ID"))
        for row in reactivated_rows
        if normalize_uid(row.get("Universal Player ID"))
    }
    unmatched_count = len(reactivated_uids - missing_uids)
    if unmatched_count == 0:
        return []
    logger.warning(
        "reconciliation_warning operation=workbook_reconciliation warning=reactivated_uid_not_subset reactivated_uid_count=%s unmatched_uid_count=%s result=warning",
        len(reactivated_uids),
        unmatched_count,
    )
    return ["Reactivated UIDs are not all present in the missing-UID handoff set"]


def _find_grand_total_row(worksheet: Worksheet | ReadOnlyWorksheet) -> int:
    for row_index in range(1, worksheet.max_row + 1):
        if worksheet.cell(row_index, 2).value == "Grand Total":
            return row_index
    raise ReconciliationError("Summary Grand Total row is missing.")


def _nonblank_uid_count(rows: Sequence[Mapping[str, object]]) -> int:
    return sum(1 for row in rows if normalize_uid(row.get("Universal Player ID")))


def _numeric(value: object) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def _round_to_cents(value: object) -> int:
    return round(_numeric(value) * 100)


def _format_cents(cents: int) -> str:
    return f"{cents / 100:.2f}"
