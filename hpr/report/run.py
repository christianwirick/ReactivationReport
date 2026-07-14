"""Run UID handoffs and build report workbooks."""

from __future__ import annotations

import logging
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from time import perf_counter

from openpyxl import Workbook

from ..diff_uid import current_uids_from_rows, missing_uids
from ..errors import (
    ExcelAutomationError,
    InputValidationError,
)
from ..excel.build_pivots import PivotAutomationResult, create_native_summary_pivot
from ..excel.read_workbook import extract_theme_xml, read_uids_from_workbook_candidates
from ..match_market import infer_market_from_rows, normalize_market_name, sheet_candidates_for_market
from ..read_tableau import read_hosted_players_csv, read_reactivated_players_csv
from . import publish
from .build import (
    add_hosted_report_sheets,
    add_reactivated_report_sheets,
    report_file_name,
)

logger = logging.getLogger(__name__)
ProgressCallback = Callable[[str], None]
REPORT_DATE_ERROR = "Report date must be M/D/YYYY, M.D.YYYY, or YYYY-MM-DD."
REPORT_DATE_FORMATS = ("%m/%d/%Y", "%m.%d.%Y", "%Y-%m-%d")


@dataclass(frozen=True)
class PreparedContext:
    """Store validated inputs shared by both workflows."""

    hosted_rows: list[dict[str, object]]
    uid_comparison: UidComparisonResult
    market: str
    report_date: date
    workbook_path: Path


@dataclass(frozen=True)
class UidComparisonResult:
    """Store current, prior and missing UID counts."""

    prior_row_count: int
    current_row_count: int
    missing_uids: tuple[str, ...]

    @property
    def missing_row_count(self) -> int:
        return len(self.missing_uids)

    @property
    def distinct_missing_uid_count(self) -> int:
        return len(set(self.missing_uids))


@dataclass(frozen=True)
class ReportResult:
    """Describe completed report build."""

    workbook_path: Path
    missing_prior_row_count: int
    distinct_missing_uid_count: int
    clipboard_uids: tuple[str, ...]
    reactivated_player_count: int
    market: str
    report_date: date
    pivot_result: PivotAutomationResult
    reactivated_completed: bool
    theme_status: str
    reconciliation_status: str


@dataclass(frozen=True)
class UidHandoffResult:
    """Describe the missing-UID handoff."""

    missing_uids: tuple[str, ...]
    missing_prior_row_count: int
    distinct_missing_uid_count: int
    clipboard_uids: tuple[str, ...]
    market: str
    report_date: date


def run_uid_handoff(
    *,
    hosted_csv_path: str | Path,
    last_week_xlsx_path: str | Path,
    market: str | None = None,
    report_date: date | None = None,
    output_folder: str | Path,
    progress_callback: ProgressCallback | None = None,
) -> UidHandoffResult:
    """Find UIDs missing from this week's report."""

    started = perf_counter()
    context = _prepare_context(
        hosted_csv_path=hosted_csv_path,
        last_week_xlsx_path=last_week_xlsx_path,
        market=market,
        report_date=report_date,
        output_folder=output_folder,
        operation="uid_handoff",
        progress_callback=progress_callback,
    )

    comparison = context.uid_comparison
    clipboard_uids = comparison.missing_uids
    result = UidHandoffResult(
        missing_uids=comparison.missing_uids,
        missing_prior_row_count=comparison.missing_row_count,
        distinct_missing_uid_count=comparison.distinct_missing_uid_count,
        clipboard_uids=clipboard_uids,
        market=context.market,
        report_date=context.report_date,
    )
    logger.info(
        "uid_handoff_completed operation=uid_handoff market=%s report_date=%s missing_row_count=%s distinct_missing_uid_count=%s clipboard_uid_count=%s duration_ms=%s result=success",
        result.market,
        result.report_date.isoformat(),
        result.missing_prior_row_count,
        result.distinct_missing_uid_count,
        len(result.clipboard_uids),
        _duration_ms(started),
    )
    return result


def run_report(
    *,
    hosted_csv_path: str | Path,
    last_week_xlsx_path: str | Path,
    market: str | None = None,
    report_date: date | None = None,
    output_folder: str | Path,
    theme_path: str | Path | None = None,
    reactivated_csv_path: str | Path | None = None,
    overwrite: bool = False,
    create_native_pivot: bool = True,
    require_native_pivot: bool = False,
    progress_callback: ProgressCallback | None = None,
) -> ReportResult:
    """Build and publish the complete report workbook."""

    started = perf_counter()
    context = _prepare_context(
        hosted_csv_path=hosted_csv_path,
        last_week_xlsx_path=last_week_xlsx_path,
        market=market,
        report_date=report_date,
        output_folder=output_folder,
        operation="report_build",
        progress_callback=progress_callback,
    )
    hosted_rows = context.hosted_rows
    comparison = context.uid_comparison
    missing = comparison.missing_uids
    resolved_market = context.market
    resolved_report_date = context.report_date
    workbook_path = context.workbook_path

    publish.ensure_can_publish(workbook_path, overwrite=overwrite)

    reactivated_rows: list[dict[str, object]] = []
    if reactivated_csv_path:
        _emit_progress(progress_callback, "Reading Reactivated Players")
        reactivated_rows = read_reactivated_players_csv(reactivated_csv_path)

    property_ids = sorted({str(row["Property ID"]) for row in hosted_rows})
    property_sheets = property_ids if len(property_ids) > 1 else []
    expected_sheets = ["Summary", resolved_market, *property_sheets]
    if reactivated_csv_path:
        expected_sheets.extend(["Reactivated Players", "Reactivation Summary"])

    publish.cleanup_abandoned_stage_files(workbook_path)
    staged_path = publish.create_staged_workbook_path(workbook_path)

    try:
        _emit_progress(progress_callback, "Building workbook")
        workbook = Workbook()
        theme_status = _apply_theme_to_workbook(workbook, theme_path)

        add_hosted_report_sheets(
            workbook=workbook,
            market=resolved_market,
            report_date=resolved_report_date,
            hosted_rows=hosted_rows,
        )
        reactivated_player_count = len(reactivated_rows)
        if reactivated_csv_path:
            add_reactivated_report_sheets(
                workbook=workbook,
                report_date=resolved_report_date,
                reactivated_rows=reactivated_rows,
            )
        _add_copy_sheet(workbook, missing)
        _force_formula_recalculation(workbook)

        try:
            publish.save_workbook(workbook, staged_path)
        finally:
            workbook.close()

        _emit_progress(progress_callback, "Validating")
        publish.validate_saved_workbook_tabs(staged_path, expected_sheets)
        logger.info(
            "workbook_validation_completed operation=workbook_publication stage=static_validation target_path=%s result=success",
            staged_path.name,
        )
        reconciliation_status = publish.reconcile_report(
            workbook_path=staged_path,
            market=resolved_market,
            hosted_rows=hosted_rows,
            reactivated_rows=reactivated_rows if reactivated_csv_path else None,
            comparison=comparison,
        )

        if create_native_pivot:
            _emit_progress(progress_callback, "Creating PivotTables")
            pivot_result = create_native_summary_pivot(
                workbook_path=staged_path,
                market_sheet=resolved_market,
                report_date=resolved_report_date,
                include_reactivation_summary=bool(reactivated_csv_path),
                theme_path=theme_path,
            )
            if pivot_result.created and pivot_result.theme_message:
                theme_status = pivot_result.theme_message
            elif pivot_result.theme_message:
                theme_status = f"{theme_status} Excel automation status: {pivot_result.theme_message}"
            if pivot_result.created:
                logger.info(
                    "native_pivot_status operation=excel_automation result=success target_path=%s",
                    staged_path.name,
                )
            else:
                logger.warning(
                    "native_pivot_status operation=excel_automation result=degraded fallback_state=static_summary target_path=%s message=%s",
                    staged_path.name,
                    pivot_result.message,
                )
        else:
            pivot_result = PivotAutomationResult(False, "Skipped native PivotTable: disabled by CLI option.")
            logger.info("native_pivot_status operation=excel_automation result=skipped reason=disabled")

        if require_native_pivot and not pivot_result.created:
            raise ExcelAutomationError(pivot_result.message)
        if require_native_pivot:
            publish.validate_required_native_pivots(
                pivot_result=pivot_result,
                include_reactivation_summary=bool(reactivated_csv_path),
            )

        _emit_progress(progress_callback, "Validating")
        publish.validate_saved_workbook_tabs(staged_path, expected_sheets)
        logger.info(
            "workbook_validation_completed operation=workbook_publication stage=post_com_validation target_path=%s result=success",
            staged_path.name,
        )
        _emit_progress(progress_callback, "Publishing")
        publish.publish_workbook(staged_path, workbook_path, overwrite=overwrite)
    except Exception:
        publish.remove_stage_file(staged_path)
        raise

    result = ReportResult(
        workbook_path=workbook_path,
        missing_prior_row_count=comparison.missing_row_count,
        distinct_missing_uid_count=comparison.distinct_missing_uid_count,
        clipboard_uids=comparison.missing_uids,
        reactivated_player_count=reactivated_player_count,
        market=resolved_market,
        report_date=resolved_report_date,
        pivot_result=pivot_result,
        reactivated_completed=bool(reactivated_csv_path),
        theme_status=theme_status,
        reconciliation_status=reconciliation_status,
    )
    logger.info(
        "report_build_completed operation=report_build market=%s report_date=%s target_path=%s missing_row_count=%s distinct_missing_uid_count=%s row_count=%s duration_ms=%s result=success",
        result.market,
        result.report_date.isoformat(),
        result.workbook_path.name,
        result.missing_prior_row_count,
        result.distinct_missing_uid_count,
        comparison.current_row_count,
        _duration_ms(started),
    )
    return result


def _prepare_context(
    *,
    hosted_csv_path: str | Path,
    last_week_xlsx_path: str | Path,
    market: str | None,
    report_date: date | None,
    output_folder: str | Path,
    operation: str,
    progress_callback: ProgressCallback | None,
) -> PreparedContext:
    """Validate inputs and and prepare shared report data."""

    started = perf_counter()
    _emit_progress(progress_callback, "Reading Hosted Players")
    hosted_rows = read_hosted_players_csv(hosted_csv_path)
    resolved_market = normalize_market_name(market) if market else infer_market_from_rows(hosted_rows)
    if not resolved_market:
        raise InputValidationError("Market name is required.")
    resolved_report_date = report_date or date.today()

    output_dir = Path(output_folder)
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise InputValidationError(f"Could not create output folder: {output_dir}") from exc
    workbook_path = output_dir / report_file_name(resolved_market, resolved_report_date)

    _emit_progress(progress_callback, "Comparing UIDs")
    last_week_uids = read_uids_from_workbook_candidates(
        last_week_xlsx_path,
        sheet_candidates_for_market(resolved_market),
    )
    current_uids = current_uids_from_rows(hosted_rows)
    comparison = UidComparisonResult(
        prior_row_count=len(last_week_uids),
        current_row_count=len(hosted_rows),
        missing_uids=tuple(missing_uids(last_week_uids, current_uids)),
    )
    logger.info(
        "missing_uid_comparison_completed operation=%s market=%s report_date=%s missing_row_count=%s distinct_missing_uid_count=%s current_row_count=%s prior_uid_row_count=%s duration_ms=%s result=success",
        operation,
        resolved_market,
        resolved_report_date.isoformat(),
        comparison.missing_row_count,
        comparison.distinct_missing_uid_count,
        comparison.current_row_count,
        comparison.prior_row_count,
        _duration_ms(started),
    )

    return PreparedContext(
        hosted_rows=hosted_rows,
        uid_comparison=comparison,
        market=resolved_market,
        report_date=resolved_report_date,
        workbook_path=workbook_path,
    )


def parse_report_date(value: str) -> date:
    """Parse a report date, defaulting to today."""

    text = value.strip()
    if not text:
        return date.today()
    for date_format in REPORT_DATE_FORMATS:
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise InputValidationError(REPORT_DATE_ERROR)


def _apply_theme_to_workbook(workbook: Workbook, theme_path: str | Path | None) -> str:
    """Apply the workbook theme and return its status."""

    if not theme_path:
        logger.warning(
            "theme_status operation=workbook_build result=degraded fallback_state=no_theme reason=no_theme_path"
        )
        return "Workbook theme not applied: no theme path was provided."
    path = Path(theme_path)
    if not path.exists():
        logger.warning(
            "theme_status operation=workbook_build result=degraded fallback_state=no_theme source_file=%s reason=file_not_found",
            path.name,
        )
        return f"Workbook theme not applied: file not found at {path}."
    if path.suffix.lower() != ".thmx":
        logger.warning(
            "theme_status operation=workbook_build result=degraded fallback_state=no_theme source_file=%s reason=invalid_extension",
            path.name,
        )
        return f"Workbook theme not applied: expected a .thmx file, got {path}."
    try:
        theme_xml = extract_theme_xml(path)
    except InputValidationError as exc:
        logger.warning(
            "theme_status operation=workbook_build result=degraded fallback_state=no_theme source_file=%s error_type=%s",
            path.name,
            type(exc).__name__,
        )
        return f"Workbook theme failed to load: {exc}"
    if not theme_xml:
        logger.warning(
            "theme_status operation=workbook_build result=degraded fallback_state=no_theme source_file=%s reason=no_theme_xml",
            path.name,
        )
        return f"Workbook theme failed to load: no theme XML found in {path.name}."
    workbook.loaded_theme = theme_xml
    logger.info(
        "theme_status operation=workbook_build result=success source_file=%s",
        path.name,
    )
    return f"Workbook theme loaded: {path.name}"


def _add_copy_sheet(workbook: Workbook, missing_uids: Sequence[str]) -> None:
    """Add a hidden backup sheet of missing UIDs."""
    if "Copy" in workbook.sheetnames:
        del workbook["Copy"]
    worksheet = workbook.create_sheet("Copy")
    worksheet["A1"] = "Universal Player ID"
    for row_index, uid in enumerate(missing_uids, start=2):
        worksheet.cell(row=row_index, column=1, value=str(uid).strip())
    worksheet.column_dimensions["A"].width = 18
    worksheet.sheet_state = "hidden"


def _force_formula_recalculation(workbook: Workbook) -> None:
    """Recalculate formulas when Excel opens the workbook."""
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except AttributeError:
        pass


def _emit_progress(progress_callback: ProgressCallback | None, stage: str) -> None:
    if progress_callback is None:
        return
    try:
        progress_callback(stage)
    except Exception:
        logger.debug("progress_callback_failed operation=progress stage=%s", stage, exc_info=True)


def _duration_ms(started: float) -> int:
    return int((perf_counter() - started) * 1000)
